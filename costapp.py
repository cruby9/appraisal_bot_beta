import pandas as pd

def find_row_and_column(square_footage, quality_rating):
    # read the excel file into a pandas dataframe
    df = pd.read_excel("costdocs/square_foot_table.xlsx", header=2, index_col=0)
    
    print("Dataframe:")
    print(df)
    
    # find the matching column based on the square footage
    column = None
    for col in df.columns:
        left, right = col.split('-')
        print(f"Checking column {col}: {left} <= {square_footage} <= {right}")
        if int(left) <= square_footage <= int(right):
            column = col
            break
    
    # find the matching row based on the quality rating
    row = None
    for i, q in enumerate(df.index):
        print(f"Checking row {q}: {q} == {quality_rating}")
        if q == quality_rating:
            row = i
            break
    
    # print the row and column numbers
    if row is not None and column is not None:
        print(f"Row {row+4}, Column {pd.Index(df.columns).get_loc(column)+2}")
        cell_value = df.loc[quality_rating, column]
        print(f"Cell value: {cell_value}")
        
        # compute the estimated construction value
        cost = square_footage * cell_value
        print(f"Estimated construction value: ${cost:,.2f}")
    else:
        print("Matching row or column not found")
        cost = None
    
    return row, column, cost

def find_basement_un_row_and_column(square_footage, quality_rating):
    # read the excel file into a pandas dataframe
    df = pd.read_excel("costdocs/Basement_un_table.xlsx", header=2, index_col=0)
    
    print("Dataframe:")
    print(df)
    
    # find the matching column based on the square footage
    column = None
    for col in df.columns:
        left, right = col.split('-')
        print(f"Checking column {col}: {left} <= {square_footage} <= {right}")
        if int(left) <= square_footage <= int(right):
            column = col
            break
    
    # find the matching row based on the quality rating
    row = None
    for i, q in enumerate(df.index):
        print(f"Checking row {q}: {q} == {quality_rating}")
        if q == quality_rating:
            row = i
            break
    
    # print the row and column numbers
    if row is not None and column is not None:
        print(f"Row {row+4}, Column {pd.Index(df.columns).get_loc(column)+2}")
        cell_value = df.loc[quality_rating, column]
        print(f"Cell value: {cell_value}")
        
        # compute the estimated construction value
        cost = square_footage * cell_value
        print(f"Estimated construction value: ${cost:,.2f}")
    else:
        print("Matching row or column not found")
        cost = None
    
    return row, column, cost

import pandas as pd

def find_basement_fn_row_and_column(square_footage, quality_rating):
    # read the excel file into a pandas dataframe
    df = pd.read_excel("costdocs/Basement_fin_table.xlsx", header=2, index_col=0)
    
    print("Dataframe:")
    print(df)
    
    # find the matching column based on the square footage
    column = None
    for col in df.columns:
        left, right = col.split('-')
        print(f"Checking column {col}: {left} <= {square_footage} <= {right}")
        if int(left) <= square_footage <= int(right):
            column = col
            break
    
    # find the matching row based on the quality rating
    row = None
    for i, q in enumerate(df.index):
        print(f"Checking row {q}: {q} == {quality_rating}")
        if q == quality_rating:
            row = i
            break
    
    # print the row and column numbers
    if row is not None and column is not None:
        print(f"Row {row+4}, Column {pd.Index(df.columns).get_loc(column)+2}")
        cell_value = df.loc[quality_rating, column]
        print(f"Cell value: {cell_value}")
        
        # compute the estimated construction value
        cost = square_footage * cell_value
        print(f"Estimated construction value: ${cost:,.2f}")
    else:
        print("Matching row or column not found")
        cost = None
    
    return row, column, cost


import pandas as pd

def find_detached_garage_row_and_column(square_footage, quality_rating):
    # read the excel file into a pandas dataframe
    df = pd.read_excel("costdocs/Detached_garage_table.xlsx", header=2, index_col=0)
    
    print("Dataframe:")
    print(df)
    
    # find the matching column based on the square footage
    column = None
    for col in df.columns:
        left, right = col.split('-')
        print(f"Checking column {col}: {left} <= {square_footage} <= {right}")
        if int(left) <= square_footage <= int(right):
            column = col
            break
    
    # find the matching row based on the quality rating
    row = None
    for i, q in enumerate(df.index):
        print(f"Checking row {q}: {q} == {quality_rating}")
        if q == quality_rating:
            row = i
            break
    
    # print the row and column numbers
    if row is not None and column is not None:
        print(f"Row {row+4}, Column {pd.Index(df.columns).get_loc(column)+2}")
        cell_value = df.loc[quality_rating, column]
        print(f"Cell value: {cell_value}")
        
        # compute the estimated construction value
        cost = square_footage * cell_value
        print(f"Estimated construction value: ${cost:,.2f}")
    else:
        print("Matching row or column not found")
        cost = None
    
    return row, column, cost

def find_attached_garage_row_and_column(square_footage, quality_rating):
    # read the excel file into a pandas dataframe
    df = pd.read_excel("costdocs/Attached_garage_table.xlsx", header=2, index_col=0)

    print("Dataframe:")
    print(df)

    # find the matching column based on the square footage
    column = None
    for col in df.columns:
        left, right = col.split('-')
        print(f"Checking column {col}: {left} <= {square_footage} <= {right}")
        if int(left) <= square_footage <= int(right):
            column = col
            break

    # find the matching row based on the quality rating
    row = None
    for i, q in enumerate(df.index):
        print(f"Checking row {q}: {q} == {quality_rating}")
        if q == quality_rating:
            row = i
            break

    # print the row and column numbers
    if row is not None and column is not None:
        print(f"Row {row+4}, Column {pd.Index(df.columns).get_loc(column)+2}")
        cell_value = df.loc[quality_rating, column]
        print(f"Cell value: {cell_value}")

        # compute the estimated construction value
        cost = square_footage * cell_value
        print(f"Estimated construction value: ${cost:,.2f}")
    else:
        print("Matching row or column not found")
        cost = None

    return row, column, cost

import openpyxl

import openpyxl

def find_patio_cost(input_value):
    # Load the Patios_table.xlsx file
    workbook = openpyxl.load_workbook('costdocs/Patios_table.xlsx')
    sheet = workbook.active
    
    # Identify the columns for the given range of values
    columns = {}
    for column in sheet.iter_cols(min_col=2, max_col=5, min_row=2, max_row=2):
        column_header = column[0].value
        column_range = tuple(map(int, column_header.split('-')))
        columns[column_range] = column[0].column_letter
    
    # Find the column for the input value
    input_column = None
    for column_range, column_letter in columns.items():
        if column_range[0] <= input_value <= column_range[1]:
            input_column = column_letter
            break
    
    # Get the cost from the input column
    if input_column is not None:
        input_row = sheet.max_row
        cost = sheet[input_column + str(input_row)].value
        total_cost = cost * input_value
        print(f"Input value: {input_value}, Cost: {total_cost}")
    else:
        print("Invalid input value")


import pandas as pd

def get_depreciation_cost(years_old, age):
    df = pd.read_excel('costdocs/deprec_table.xlsx', index_col=0)
    print("DataFrame before lookup:")
    print(df)
    col_num = df.columns.get_loc(years_old)
    print(f"Column number: {col_num}")
    row_num = df.index.get_loc(age)
    print(f"Row label: {row}")
    depreciation_rate = df.iloc[row_num, col_num]
    #print(f"Depreciation figure: {cost}")
    cost = depreciation_rate
    return cost



if __name__ == '__main__':
    # call the function with the desired input value for finished basement
    row, col, cost = find_row_and_column(1150, 2.5)

    # print the estimated cost
    print(f"The cost for a {row} quality rated home with {col} sq.ft. is ${cost:,.2f}")

    # call the function with the desired input values for unfinished basement
    row, col, cost = find_basement_un_row_and_column(500, 2.5)

    # print the estimated cost
    print(f"The cost for a {row} quality rated home with {col} sq.ft. is ${cost:,.2f}")

    # call the function with the desired input values for detached garage
    row, col, cost = find_detached_garage_row_and_column(400, 2)

    # print the estimated cost
    print(f"The cost for a {col} sq.ft. detached garage is ${cost:,.2f}")

    # call the function with the desired input values for attached garage
    row, col, cost = find_attached_garage_row_and_column(300, 2)

    # print the estimated cost
    print(f"The cost for a {col} sq.ft. attached garage is ${cost:,.2f}")

    # call the function with the desired input value for patio
    find_patio_cost(350)
   

    # Call get_depreciation_cost function with example values
    cost = get_depreciation_cost(60, 8)
    print(f"Depreciation cost: {cost}")








